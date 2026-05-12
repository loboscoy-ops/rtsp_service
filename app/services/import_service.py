from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from app import config
from app.database.repository import Repository
from app.utils.validators import is_valid_rtsp_url, parse_enabled


REQUIRED_FIELDS = ("object_name", "camera_identifier", "rtsp_url")
OPTIONAL_FIELDS = ("camera_name", "group_name", "gps_coords", "uin", "enabled")
ALL_FIELDS = REQUIRED_FIELDS + OPTIONAL_FIELDS

# Подсказки для авто-подбора колонок (точные совпадения и частичные).
FIELD_HINTS: dict[str, tuple[str, ...]] = {
    "object_name": (
        "object_name",
        "наименование объекта",
        "объект",
        "проект",
        "площадка",
    ),
    "camera_identifier": (
        "camera_identifier",
        "№ п/п",
        "номер",
        "id камеры",
        "идентификатор камеры",
        "идентификатор",
        "id",
    ),
    "camera_name": (
        "camera_name",
        "имя камеры",
        "наименование камеры",
        "описание зоны обзора камеры",
        "описание зоны",
        "имя камеры в локальной системе видеонаблюдения",
    ),
    "rtsp_url": (
        "rtsp_url",
        "ссылка на видеотрансляцию",
        "rtsp",
        "url",
        "ссылка",
        "поток",
    ),
    "group_name": (
        "group_name",
        "группа",
        "зона",
        "тип камеры",
        "место установки камеры",
    ),
    "gps_coords": (
        "gps_coords",
        "gps координаты",
        "gps",
        "координаты",
        "координата",
    ),
    "uin": (
        "uin",
        "уин",
        "код объекта",
    ),
    "enabled": (
        "enabled",
        "активна",
        "вкл",
        "включена",
    ),
}


@dataclass
class PreviewIssue:
    row_number: int
    message: str


@dataclass
class PreviewRow:
    object_name: str
    camera_identifier: str
    camera_name: str
    rtsp_url: str
    group_name: str
    gps_coords: str
    uin: str
    enabled: bool
    valid: bool
    error: str = ""


@dataclass
class ImportPreview:
    rows: list[PreviewRow] = field(default_factory=list)
    issues: list[PreviewIssue] = field(default_factory=list)

    @property
    def valid_rows(self) -> list[PreviewRow]:
        return [r for r in self.rows if r.valid]


@dataclass
class SheetData:
    name: str
    rows: list[list[str]]


def _norm(text: str) -> str:
    return " ".join((text or "").strip().lower().replace("ё", "е").split())


class ImportService:
    def __init__(self, repository: Repository):
        self.repository = repository

    @staticmethod
    def _clean_cell(value: object) -> str:
        """Быстрая нормализация ячейки без обращений к pandas.

        ``openpyxl`` уже отдаёт ``None`` для пустых ячеек и числа/datetime —
        в native-типах. Дергать ``pd.isna`` на каждую ячейку дорого, поэтому
        импорт раньше «жил» долго на больших листах.
        """
        if value is None:
            return ""
        if isinstance(value, float):
            # NaN — единственное float, которое != самому себе.
            if value != value:  # noqa: PLR0124
                return ""
            if value.is_integer():
                return str(int(value)).strip()
            return repr(value).strip()
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def list_sheet_names(self, path: Path) -> list[str]:
        """Лёгкая операция: получить только имена листов без чтения данных."""
        suffix = path.suffix.lower()
        if suffix == ".xls":
            import xlrd  # type: ignore

            book = xlrd.open_workbook(str(path), on_demand=True)
            try:
                return list(book.sheet_names())
            finally:
                book.release_resources()
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def read_sheet(self, path: Path, name: str) -> SheetData:
        """Прочитать содержимое одного листа максимально дёшево."""
        suffix = path.suffix.lower()
        clean = self._clean_cell
        if suffix == ".xls":
            import xlrd  # type: ignore

            book = xlrd.open_workbook(str(path), on_demand=True)
            try:
                sheet = book.sheet_by_name(name)
                rows: list[list[str]] = []
                for r in range(sheet.nrows):
                    row = [clean(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                    rows.append(row)
            finally:
                book.release_resources()
        else:
            from openpyxl import load_workbook

            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            try:
                sheet = wb[name]
                rows = []
                for raw_row in sheet.iter_rows(values_only=True):
                    rows.append([clean(c) for c in raw_row])
            finally:
                wb.close()

        # Срезаем пустой «хвост» (типичная проблема Excel-файлов с лишними строками).
        while rows and not any(c for c in rows[-1]):
            rows.pop()
        return SheetData(name=name, rows=rows)

    def list_sheets(self, path: Path) -> list[SheetData]:
        """Совместимое API: читает все листы.

        Используйте :meth:`list_sheet_names` + :meth:`read_sheet` для отложенной
        загрузки — это намного быстрее для больших книг.
        """
        return [self.read_sheet(path, name) for name in self.list_sheet_names(path)]

    def refine_identifier_mapping(
        self,
        sheet: SheetData,
        header_row_index: int,
        mapping: dict[str, int | None],
    ) -> dict[str, int | None]:
        """Если в выбранной колонке camera_identifier данные плохо уникальны,
        попробовать другие подходящие колонки и выбрать ту, что даёт больше
        уникальных значений в строках, где есть валидный rtsp_url."""
        url_col = mapping.get("rtsp_url")
        if url_col is None:
            return mapping
        # отбираем строки, в которых rtsp валиден
        data_rows: list[list[str]] = []
        for r_idx in range(header_row_index + 1, len(sheet.rows)):
            row = sheet.rows[r_idx]
            if url_col >= len(row):
                continue
            if is_valid_rtsp_url(row[url_col]):
                data_rows.append(row)
        if not data_rows:
            return mapping

        candidate_cols = sorted(
            {c for c in (mapping.get("camera_identifier"), *self._identifier_candidate_cols(sheet, header_row_index)) if c is not None}
        )
        used = {v for k, v in mapping.items() if v is not None and k != "camera_identifier"}
        best_col = mapping.get("camera_identifier")
        best_unique = self._count_unique(data_rows, best_col) if best_col is not None else 0
        for col in candidate_cols:
            if col in used:
                continue
            unique = self._count_unique(data_rows, col)
            if unique > best_unique:
                best_unique = unique
                best_col = col
        if best_col is not None and best_col != mapping.get("camera_identifier"):
            mapping = dict(mapping)
            mapping["camera_identifier"] = best_col
        return mapping

    @staticmethod
    def _count_unique(rows: list[list[str]], col: int) -> int:
        if col is None:
            return 0
        seen: set[str] = set()
        for row in rows:
            if col >= len(row):
                continue
            v = row[col].strip()
            if v:
                seen.add(v.lower())
        return len(seen)

    def _identifier_candidate_cols(self, sheet: SheetData, header_row_index: int) -> list[int]:
        """Колонки, претендующие на роль ID: содержат числа или подходящие подписи."""
        if header_row_index >= len(sheet.rows):
            return []
        header = sheet.rows[header_row_index]
        hints = {_norm(h) for h in FIELD_HINTS["camera_identifier"]}
        cols: list[int] = []
        for idx, h in enumerate(header):
            n = _norm(h)
            if not n:
                continue
            if any(hint in n for hint in hints):
                cols.append(idx)
        return cols

    def detect_header_row(self, rows: list[list[str]]) -> int:
        """Подбирает строку заголовков, в которой максимум совпадений с FIELD_HINTS."""
        best_idx = 0
        best_score = -1
        best_nonempty = -1
        limit = min(len(rows), 30)
        for idx in range(limit):
            row = rows[idx]
            mapping = self.auto_detect_mapping(row)
            score = sum(1 for v in mapping.values() if v is not None)
            non_empty = sum(1 for c in row if c)
            if score > best_score or (score == best_score and non_empty > best_nonempty):
                best_score = score
                best_nonempty = non_empty
                best_idx = idx
        return best_idx

    def build_synthetic_headers(
        self, sheet: SheetData, header_row_index: int
    ) -> list[str]:
        """Объединяет подписи из header_row и всех строк выше — для случаев,
        когда заголовки разнесены по нескольким объединённым строкам."""
        if not sheet.rows or header_row_index < 0:
            return []
        max_cols = max((len(r) for r in sheet.rows), default=0)
        out: list[str] = []
        for col in range(max_cols):
            parts: list[str] = []
            for r in range(0, min(header_row_index + 1, len(sheet.rows))):
                row = sheet.rows[r]
                if col < len(row):
                    v = " ".join(row[col].split())
                    if v and v not in parts:
                        parts.append(v)
            out.append(" / ".join(parts))
        return out

    def auto_detect_mapping(self, header_row: list[str]) -> dict[str, int | None]:
        normalized = [_norm(c) for c in header_row]
        used: set[int] = set()
        mapping: dict[str, int | None] = {f: None for f in ALL_FIELDS}

        # Для поля "Группа/тип" всегда предпочитаем колонку "Типы камер",
        # если она присутствует в заголовках формы.
        preferred_group_col = self._preferred_group_name_col(normalized)
        if preferred_group_col is not None:
            mapping["group_name"] = preferred_group_col
            used.add(preferred_group_col)

        # 1. exact match first
        for field_name, hints in FIELD_HINTS.items():
            if mapping[field_name] is not None:
                continue
            wanted = {_norm(h) for h in hints}
            for idx, h in enumerate(normalized):
                if not h or idx in used:
                    continue
                if h in wanted:
                    mapping[field_name] = idx
                    used.add(idx)
                    break
        # 2. fuzzy substring fallback for fields not yet mapped
        for field_name, hints in FIELD_HINTS.items():
            if mapping[field_name] is not None:
                continue
            for idx, h in enumerate(normalized):
                if not h or idx in used:
                    continue
                if any(_norm(hint) in h for hint in hints):
                    mapping[field_name] = idx
                    used.add(idx)
                    break
        return mapping

    @staticmethod
    def _preferred_group_name_col(normalized_header: list[str]) -> int | None:
        preferred_tokens = ("типы камер", "типы камеры")
        for idx, title in enumerate(normalized_header):
            if not title:
                continue
            if any(token in title for token in preferred_tokens):
                return idx
        return None

    def build_preview_from_mapping(
        self,
        sheet: SheetData,
        header_row_index: int,
        mapping: dict[str, int | None],
    ) -> ImportPreview:
        preview = ImportPreview()
        if not sheet.rows:
            preview.issues.append(PreviewIssue(0, "Пустой лист"))
            return preview

        for required in REQUIRED_FIELDS:
            col_idx = mapping.get(required)
            if col_idx is None:
                preview.issues.append(
                    PreviewIssue(
                        0,
                        f"Не задано соответствие колонки для поля «{required}»",
                    )
                )
        if preview.issues:
            return preview

        # Распространяем значения из объединённых ячеек: если object_name пуст,
        # подставляем последнее непустое значение из той же колонки выше.
        last_object_name = ""
        seen_urls: dict[str, int] = {}

        for r_idx in range(header_row_index + 1, len(sheet.rows)):
            row = sheet.rows[r_idx]
            row_no = r_idx + 1

            def _val(field_name: str) -> str:
                col = mapping.get(field_name)
                if col is None or col < 0 or col >= len(row):
                    return ""
                return row[col]

            rtsp_url = _val("rtsp_url")
            # Тихо пропускаем строки без RTSP — это служебные/нумерационные/пустые строки.
            if not is_valid_rtsp_url(rtsp_url):
                continue

            object_name = _val("object_name") or last_object_name
            if _val("object_name"):
                last_object_name = _val("object_name")
            camera_identifier = _val("camera_identifier")
            camera_name = _val("camera_name")
            group_name = _val("group_name")
            gps_coords = _val("gps_coords")
            uin = _val("uin")
            enabled_raw = _val("enabled")
            enabled = parse_enabled(enabled_raw) if enabled_raw else True

            if not camera_name:
                camera_name = (
                    f"Камера {camera_identifier}" if camera_identifier else f"Строка {row_no}"
                )

            err_parts: list[str] = []
            if not object_name:
                err_parts.append("object_name пустой")
            if not camera_identifier:
                err_parts.append("camera_identifier пустой")

            url_key = rtsp_url.strip().lower()
            if url_key:
                if url_key in seen_urls:
                    err_parts.append(
                        f"дубликат RTSP (см. строку {seen_urls[url_key]})"
                    )
                else:
                    seen_urls[url_key] = row_no

            err = "; ".join(err_parts)
            preview.rows.append(
                PreviewRow(
                    object_name=object_name,
                    camera_identifier=camera_identifier,
                    camera_name=camera_name,
                    rtsp_url=rtsp_url,
                    group_name=group_name,
                    gps_coords=gps_coords,
                    uin=uin,
                    enabled=enabled,
                    valid=not err_parts,
                    error=err,
                )
            )
            if err_parts:
                preview.issues.append(PreviewIssue(row_no, err))

        # Если внутри объекта несколько строк имеют одинаковый camera_identifier,
        # но разные RTSP (это разные камеры), добавляем суффикс '-2', '-3', ...
        seen_pairs: dict[tuple[str, str], int] = {}
        for row in preview.rows:
            if not row.camera_identifier:
                continue
            key = (row.object_name.lower(), row.camera_identifier.lower())
            count = seen_pairs.get(key, 0) + 1
            seen_pairs[key] = count
            if count > 1:
                suffixed = f"{row.camera_identifier}-{count}"
                preview.issues.append(
                    PreviewIssue(
                        0,
                        f"В объекте «{row.object_name}» уже был camera_identifier "
                        f"'{row.camera_identifier}' — переименовано в '{suffixed}'.",
                    )
                )
                row.camera_identifier = suffixed

        return preview

    def import_valid_rows(self, preview: ImportPreview) -> tuple[int, int]:
        valid = preview.valid_rows
        if not valid:
            return 0, 0
        rows_payload: Iterable[dict] = (
            {
                "object_name": row.object_name,
                "camera_identifier": row.camera_identifier,
                "camera_name": row.camera_name,
                "group_name": row.group_name,
                "rtsp_url": row.rtsp_url,
                "enabled": row.enabled,
                "gps_coords": row.gps_coords,
                "uin": row.uin,
            }
            for row in valid
        )
        return self.repository.bulk_upsert_cameras(rows_payload)
