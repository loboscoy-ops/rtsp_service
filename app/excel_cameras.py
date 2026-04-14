from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import load_workbook

from .models import CameraRecord
from .rtsp_probe import looks_like_rtsp_url

log = logging.getLogger(__name__)

_FLOAT_RE = re.compile(r"-?\d+(?:[.,]\d+)?")


def parse_lat_lon(raw: object) -> tuple[float | None, float | None]:
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s:
        return None, None
    nums: list[float] = []
    for m in _FLOAT_RE.finditer(s.replace(",", ".")):
        try:
            nums.append(float(m.group(0)))
        except ValueError:
            continue
    if len(nums) < 2:
        return None, None
    a, b = nums[0], nums[1]
    if -90 <= a <= 90 and -180 <= b <= 180:
        return a, b
    if -90 <= b <= 90 and -180 <= a <= 180:
        return b, a
    return None, None


# Порядок важен: «адрес объекта» раньше, чем url, чтобы не перепутать с колонкой ссылки.
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "row_no": ("№ п/п", "№п/п", "п/п", "номер по порядку", "№"),
    "uin": ("уин", "uin"),
    "name": (
        "наименование объекта",
        "наименование камеры",
        "наименование",
        "название",
        "камера",
        "name",
        "точка",
        "имя",
    ),
    "address": ("адрес объекта", "адрес"),
    "coord": (
        "gps координаты",
        "gps координат",
        "gps",
        "координата",
        "координаты",
        "geo",
        "широта",
        "location",
        "map",
    ),
    "url": (
        "ссылка на видеотрансляцию",
        "видеотрансляцию",
        "видеотрансляция",
        "трансляция",
        "rtsp",
        "rtsp link",
        "rtsp url",
        "url",
        "поток",
        "link",
        "стрим",
    ),
    "camera_type": (
        "тип камеры",
        "тип оборудования",
        "тип",
        "type",
        "модель",
        "model",
    ),
    "project": (
        "проект",
        "project",
        "объект",
        "площадка",
        "назначение",
        "вид работ",
        "работы",
    ),
}


def _norm_header(s: str) -> str:
    t = (s or "").strip().lower().replace("ё", "е")
    return " ".join(t.split())


def _col_index_to_a1(idx: int) -> str:
    n = idx + 1
    letters = []
    while n:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(65 + rem))
    return "".join(reversed(letters))


def _map_header_row(header_cells: list[str]) -> dict[str, int] | None:
    col_map: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        key = _norm_header(cell)
        if not key:
            continue
        for field_name, aliases in HEADER_ALIASES.items():
            if field_name in col_map:
                continue
            if any(key == a or key.startswith(a + " ") for a in aliases):
                col_map[field_name] = idx
                break
    if "url" not in col_map:
        return None
    for k in HEADER_ALIASES:
        if k not in col_map:
            col_map[k] = -1

    uc = col_map.get("url", -1)
    if uc >= 0:
        if col_map.get("camera_type", -1) == -1:
            col_map["camera_type"] = uc + 1
        if col_map.get("project", -1) == -1:
            col_map["project"] = uc + 2

    return col_map


def _cell(row: list[str], idx: int) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return (row[idx] or "").strip()


def _normalize_row(row: tuple) -> list[str]:
    out: list[str] = []
    for c in row:
        if c is None:
            out.append("")
        else:
            out.append(str(c).strip())
    return out


def load_cameras_from_excel(path: Path) -> tuple[list[CameraRecord], str | None]:
    if not path.is_file():
        return [], f"Файл не найден: {path} (загрузите .xlsx через форму)"

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            raw_rows = list(
                ws.iter_rows(min_row=1, max_row=500, max_col=30, values_only=True)
            )
        finally:
            wb.close()
    except Exception as e:
        log.exception("excel open")
        return [], f"Не удалось открыть Excel: {e}"

    if not raw_rows:
        return [], "Пустой лист"

    rows = [_normalize_row(r) for r in raw_rows]
    while rows and not any(x for x in rows[-1]):
        rows.pop()
    if not rows:
        return [], "Пустой лист"

    col_map = _map_header_row(rows[0])
    if not col_map:
        col_map = {
            "row_no": 0,
            "uin": 1,
            "name": 2,
            "address": 3,
            "coord": 4,
            "url": 5,
            "camera_type": 6,
            "project": 7,
        }
        log.info(
            "заголовки не распознаны: A=№ B=УИН C=наименование D=адрес E=GPS F=ссылка G=тип H=проект"
        )

    max_ci = max(col_map.values()) if col_map else 0

    cameras: list[CameraRecord] = []
    for r_idx in range(1, len(rows)):
        row = list(rows[r_idx])
        while len(row) <= max_ci:
            row.append("")
        url = _cell(row, col_map["url"])
        if not looks_like_rtsp_url(url):
            continue
        excel_row = r_idx + 1
        pj = _cell(row, col_map["project"])
        nm = _cell(row, col_map["name"])
        ct = _cell(row, col_map["camera_type"])
        coord_raw = _cell(row, col_map["coord"]) if col_map.get("coord", -1) >= 0 else ""
        lat, lon = parse_lat_lon(coord_raw)
        rn = _cell(row, col_map["row_no"]) if col_map.get("row_no", -1) >= 0 else ""
        uin = _cell(row, col_map["uin"]) if col_map.get("uin", -1) >= 0 else ""
        addr = _cell(row, col_map["address"]) if col_map.get("address", -1) >= 0 else ""
        if not nm:
            nm = f"Строка {excel_row}"
        cid = f"E_{excel_row}"
        ucol = col_map["url"]
        cell_a1 = f"{_col_index_to_a1(ucol)}{excel_row}"
        cameras.append(
            CameraRecord(
                camera_id=cid,
                rtsp_url=url,
                project=pj,
                name=nm,
                camera_type=ct,
                lat=lat,
                lon=lon,
                row_no=rn,
                uin=uin,
                address=addr,
                cell_a1=cell_a1,
            )
        )

    cameras.sort(key=lambda c: (c.project.lower(), c.name.lower()))
    return cameras, None
